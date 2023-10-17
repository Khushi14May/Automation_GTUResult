import openpyxl 
from PIL import Image
import pyperclip
from openpyxl.styles import PatternFill
import pandas as pd
from browsermobproxy import Server
from selenium import webdriver

excel_file_path="C:/Users/158421/OneDrive - Arrow Electronics, Inc/Desktop/GTU-result-Robot/Automation_GTUResult/output/Result-Robot.xlsx"
def crop_image(input_path, output_path, left, top, right, bottom):
    try:
        img = Image.open(input_path)
        img_width, img_height = img.size
        left = max(0, left)
        top = max(0, top)
        right = min(img_width, right)
        bottom = min(img_height, bottom)
        cropped_img = img.crop((left, top, right, bottom))
        cropped_img.save(output_path)
        print(f"Image cropped and saved.")
        return  True
    except Exception as e:
        print(f"Error: {e}")
        return  False

def get_text_from_clipboard():
    clipboard_value = pyperclip.paste()
    return clipboard_value

def enter_data_excel(list_data):
    global  excel_file_path
    book = openpyxl.load_workbook(excel_file_path, read_only=False)
    sheet = book.active
    row_count=sheet.max_row
    col_count=sheet.max_column
    for i in range(1, col_count + 1):  # to get rows
        sheet.cell(row=row_count+1, column=i, value=list_data[i-1])
        book.save(excel_file_path)
            
def fill_color():
    book = openpyxl.load_workbook(excel_file_path, read_only=False)
    sheet = book.active
    row_count=sheet.max_row
    for i in range(2,row_count+1):
        cell = sheet.cell(row=i, column=10) 
        if cell.value=="Pass":
            fill = PatternFill(start_color="FF00FF00", end_color="FF00FF00", fill_type="solid")
            cell.fill = fill
        elif cell.value=="Fail":
            fill =PatternFill(start_color="FFFF0000", end_color="FFFF0000", fill_type="solid")
            cell.fill = fill
        else:
            fill =PatternFill(start_color="FFFFFF00", end_color="FFFFFF00", fill_type="solid")
            cell.fill = fill
        book.save(excel_file_path)
        
def sort_excel():
    global  excel_file_path
    sheet_name = 'Sheet1'
    df = pd.read_excel(excel_file_path, sheet_name=sheet_name)
    column_to_sort_by = 'CGPA'
    df.sort_values(by=column_to_sort_by, inplace=True, ascending=False,key=lambda x: x.replace('_ _', '0'))
    df.to_excel(excel_file_path, sheet_name=sheet_name, index=False)


